import os
from openpyxl import load_workbook

def batch_replace_in_excel_folder(folder_path, old_text, new_text):
    for root, dirs, files in os.walk(folder_path):
        for filename in files:
            if filename.endswith(".xlsx"):
                file_path = os.path.join(root, filename)

                wb = load_workbook(file_path)

                for sheetname in wb.sheetnames:
                    ws = wb[sheetname]

                    for row in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column):
                        for cell in row:
                            if cell.value and old_text in str(cell.value):
                                cell.value = str(cell.value).replace(old_text, new_text)

                wb.save(file_path)
                print(f"成功替换 '{old_text}' 为 '{new_text}' 在文件: {filename}")

folder_path = r""
old_text = ""
new_text = ""

batch_replace_in_excel_folder(folder_path, old_text, new_text)